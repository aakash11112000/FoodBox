#Case Study-Food box
#Done By
#Syntax error team

from tkinter import *
from openpyxl import *
import os
from datetime import *
import string
import tkinter.messagebox

#__________________________________________________________________________________________________________________________________
#----------------------------------------Beginning of the inventory module class--------------------------------------------------------------------------------

class Inventory():
    def __init__(self):
        self.number= 0
        self.value1= 0
        self.rowrow= 4
        self.color=['blue','red','orange','green','navy','firebrick','cyan','turquoise']
           
    def inventory(self):
        inv =Toplevel()
        inv.title('The Food Box - INVENTORY')
        inv.configure(bg='skyblue1')
        wb=load_workbook('master_sheet.xlsx')
        ws=wb['Inventory']; row=ws.max_row
        ws1=wb['Sales']
        ws2=wb['Report']
        now=datetime.now()
        today=str(now.day)+'/'+str(now.month)+'/'+str(now.year)
        Label(inv,text=today,bg='skyblue1',fg='black',font=('Times',20,'bold')).grid(row=0,column=2,sticky=N+E+W+S)
        img1=PhotoImage(file='csfb_logo.gif')
        Label(inv,image=img1).grid(row=0,column=6)
        for i in range(1,row+1):
            s1='a'+str(i);a1=str(ws[s1].value)
            s2='b'+str(i);a2=str(ws[s2].value)
            s3='c'+str(i);a3=str(ws[s3].value)
            s4='d'+str(i);a4=str(ws[s4].value)
            s5='e'+str(i);a5=str(ws[s5].value)
            if a1!='None' and a2!='None' and a3!='None' and a4!='None' and a5!='None':
                Label(inv,text=a1.upper(),bg='skyblue1',fg='black',font=('Times',15,'bold')).grid(row=i,column=0,sticky=W)
                Label(inv,text=a2.upper(),bg='skyblue1',fg='black',font=('Times',15,'bold')).grid(row=i,column=1,sticky=W)
                Label(inv,text=a3.capitalize(),bg='skyblue1',fg='black',font=('Times',15,'bold')).grid(row=i,column=2,sticky=W)
                Label(inv,text=a4,bg='skyblue1',fg='black',font=('Times',15,'bold')).grid(row=i,column=3,sticky=W)
                Label(inv,text='Rs. '+a5,bg='skyblue1',fg='black',font=('Times',15,'bold')).grid(row=i,column=4,sticky=W)
                if i>1:
                    Button(inv,text='MODIFY FOOD',bg='yellow',fg='black',font=('Times',15),command=lambda num1=i: self.modify_food(wb,ws,ws1,ws2,inv,num1)).grid(row=i,column=5,padx=5)
            num=i
        Button(inv,text='ADD FOOD',bg='yellow',fg='black',font=('Times',15),command=lambda: self.add_new(wb,ws,ws1,ws2,inv)).grid(row=num+1,column=0,padx=5)
        
        Button(inv,text='DELETE FOOD',bg='yellow',fg='black',font=('Times',15),command=lambda: self.delete(wb,ws,ws1,ws2,inv)).grid(row=num+1,column=1,padx=5)
        Button(inv,text='SEARCH BY NAME',bg='yellow',fg='black',font=('Times',15),command=lambda: self.search_by_name(wb,ws,ws1,ws2)).grid(row=num+1,column=2,padx=5)
        Button(inv,text='SEARCH BY FOOD NUMBER',bg='yellow',fg='black',font=('Times',15),command=lambda: self.search_by_fno(wb,ws,ws1,ws2)).grid(row=num+1,column=3,padx=5)
        Button(inv,text='GENERATE REPORT',bg='yellow',fg='black',font=('Times',15),command=lambda: self.generate_report(wb,ws,ws1,ws2)).grid(row=num+2,column=2,padx=5)
        
        inv.mainloop()
        return inv
#_______________________________________________________________________________________________________________________________________________________________________
#____________________________________________Function to add new foods to the inventory______________________________________________________________________________________    
            
    def add_new(self,wb,ws,ws1,ws2,inv):            
        def b1():
            flag=True
            maxrow=ws.max_row
            foodtype=e1.get()
            foodno=e2.get()
            foodname=e3.get()
            qty=e4.get()
            price=e5.get()
            if foodtype=='' or foodno=='' or foodname=='' or qty=='' or price=='' :
                tkinter.messagebox.showerror('Error','All fields are mandatory!')
            else:
                for i in range(2,maxrow+1):
                    if str(foodno).upper() == str(ws['b'+str(i)].value).upper():                  
                        tkinter.messagebox.showerror('Error','Food Number already Exists. \n Please choose a new one')
                        flag=False     
                if flag:
                    if foodtype.upper() == "CHINESE" or foodtype.upper() == "CONTINENTAL" or foodtype.upper() == "SOUTH" or foodtype.upper() == "NORTH":
                        try:
                            if int(qty)>0 and int(price)>=0:
                                foodtype=foodtype.upper()
                                foodno=foodno.upper()
                                foodname=foodname.capitalize()
                                ws.append([foodtype,foodno,foodname,qty,price])
                                ws1.append([foodtype,foodno,foodname,'0',price])
                                ws2.append([foodtype,foodno,foodname,'0',price])
                                wb.save("master_sheet1.xlsx")
                                os.remove("master_sheet.xlsx")
                                os.rename("master_sheet1.xlsx","master_sheet.xlsx")
                                tkinter.messagebox.showinfo('Success','New Food added')
                                add.destroy()
                                inv.destroy()
                                self.inventory()
                            elif qty<=0:
                                tkinter.messagebox.showerror('Error','Quantity must be a positive non-zero integer')
                        except:
                            tkinter.messagebox.showerror('Error','Wrong Data for Quantity or Price')
                    else:
                            tkinter.messagebox.showerror('Error','Please enter a valid foodtype')
                        
        add=Toplevel()
        add.configure(bg='RoyalBlue1')
        add.title('The Food Box - Add New Food')
        e1=StringVar(); e2=StringVar(); e3=StringVar(); e4=StringVar(); e5=StringVar()     
        Label(add,text="Enter the details",font=('Times',20))
        Label(add,text="Food Type",font=('Times',20),bg='RoyalBlue1').grid(row=0,column=0,sticky=W,pady=3)
        Label(add,text="Food Number",font=('Times',20),bg='RoyalBlue1').grid(row=1,column=0,sticky=W,pady=3)
        Label(add,text="Food Name",font=('Times',20),bg='RoyalBlue1').grid(row=2,column=0,sticky=W,pady=3)
        Label(add,text="Quantity",font=('Times',20),bg='RoyalBlue1').grid(row=3,column=0,sticky=W,pady=3)
        Label(add,text="Price Rs. ",font=('Times',20),bg='RoyalBlue1').grid(row=4,column=0,sticky=W,pady=3)
        Entry(add,textvariable=e1,font=('Times',20)).grid(row=0,column=1,sticky=W,pady=3)
        Entry(add,textvariable=e2,font=('Times',20)).grid(row=1,column=1,sticky=W,pady=3)
        Entry(add,textvariable=e3,font=('Times',20)).grid(row=2,column=1,sticky=W,pady=3)
        Entry(add,textvariable=e4,font=('Times',20)).grid(row=3,column=1,sticky=W,pady=3)
        Entry(add,textvariable=e5,font=('Times',20)).grid(row=4,column=1,sticky=W,pady=3)
        Button(add,text="ADD FOOD",font=('Times',25),fg='blue2',bg='green',command=b1).grid(row=5,column=2,sticky=W,padx=5,pady=5)
        add.mainloop()
       
#_____________________________________________________________________________________________________________________________________________
#____________________________________________Function to modify existing foods in the inventory__________________________________________________________
       
    def modify_food(self,wb,ws,ws1,ws2,inv,num1):
        def modify(somelist):
            foodtype=e1.get()
            foodno=e2.get()
            foodname=e3.get()
            qty=e4.get()
            price=e5.get()          
            maxrow=ws.max_row
            if (foodtype.upper()==somelist[0]) and (foodno.upper()==somelist[1]) and (foodname  ==somelist[2]) and (qty==somelist[3]) and (price==somelist[4]):
                tkinter.messagebox.showinfo('Notification','No changes have been made')
                modify_main.destroy()
            else:
                if foodtype=='' or foodno=='' or foodname=='' or qty=='' or price=='':
                    tkinter.messagebox.showerror('Error','All fields are mandatory')
                if foodtype.upper() == "SOUTH" or foodtype.upper() == "NORTH" or foodtype.upper() == "CONTINENTAL" or foodtype.upper() == "CHINESE":
                    if str(qty).isdigit() or qty=='' :
                        if price=='' or str(price).isdigit():
                            if qty>=0 and price>=0:
                                for i in range(2,maxrow+1):
                                    if foodtype!='':
                                        ws['a'+self.number]=foodtype
                                        ws1['a'+self.number]=foodtype
                                        ws2['a'+self.number]=foodtype
                                    if foodno!='':
                                        ws['b'+self.number]=foodno
                                        ws1['b'+self.number]=foodno
                                        ws2['b'+self.number]=foodno
                                       
                                    if foodname!='':
                                        ws['c'+self.number]=foodname
                                        ws1['c'+self.number]=foodname
                                        ws2['c'+self.number]=foodname
                                    if qty!='':
                                        ws['d'+self.number]=qty
                                    if price!='':
                                        ws['e'+self.number]=price
                                        ws1['e'+self.number]=price
                                        ws2['e'+self.number]=price
                                wb.save("master_sheet.xlsx")
                                if foodtype=='' and foodno=='' and foodname=='' and qty=='' and price=='':
                                    tkinter.messagebox.showinfo('Notification','No changes have been made')
                                else:
                                    tkinter.messagebox.showinfo('SUCCESS!','Food Item Modified')
                                    modify_main.destroy()
                                    inv.destroy()
                                    self.inventory()

                            else:
                                tkinter.messagebox.showerror('Error',"Negative values for field not allowed")
                        else:
                            tkinter.messagebox.showerror('Error',"Only positive numerical characters are allowed for price")
                    else:
                        tkinter.messagebox.showerror('Error',"Only positive numerical characters are allowed for quantity")
                else:
                    tkinter.messagebox.showerror('Error',"Enter a valid foodtype \n NORTH/SOUTH/CONTINENTAL/CHINESE")

        modify_main=Toplevel()
        modify_main.configure(bg='purple2')
        self.number = str(num1)
        modify_main.title('The Food Box-Modify Existing Food Item')
        e1=StringVar(value=ws['a'+str(self.number)].value)
        e2=StringVar(value=ws['b'+str(self.number)].value)
        e3=StringVar(value=ws['c'+str(self.number)].value)           
        e4=StringVar(value=ws['d'+str(self.number)].value)
        e5=StringVar(value=ws['e'+str(self.number)].value)
        somelist=[ws['a'+str(self.number)].value,ws['b'+str(self.number)].value,ws['c'+str(self.number)].value,ws['d'+str(self.number)].value,ws['e'+str(self.number)].value]
        maxrow=ws.max_row
        j=0
        for i in range(1,num1+1,num1-1):
           
            s1='a'+str(i);a1=str(ws[s1].value)
            s2='b'+str(i);a2=str(ws[s2].value)
            s3='c'+str(i);a3=str(ws[s3].value)
            s4='d'+str(i);a4=str(ws[s4].value)
            s5='e'+str(i);a5=str(ws[s5].value)

            Label(modify_main,text=a1.upper(),fg='black',font=('Times',15,'bold'),bg='purple2').grid(row=j,column=0,sticky=W)
            Label(modify_main,text=a2.upper(),fg='black',font=('Times',15,'bold'),bg='purple2').grid(row=j,column=1,sticky=W)
            Label(modify_main,text=a3,fg='black',font=('Times',15,'bold'),bg='purple2').grid(row=j,column=2,sticky=W)
            Label(modify_main,text=a4,fg='black',font=('Times',15,'bold'),bg='purple2').grid(row=j,column=3,sticky=W)
            Label(modify_main,text=a5,fg='black',font=('Times',15,'bold'),bg='purple2').grid(row=j,column=4,sticky=W)
            j+=1

        Label(modify_main,text="New Food Type: ",font=('Times',20),bg='purple2').grid(row=3,column=0,sticky=W)
        Label(modify_main,text="New Food Number: ",font=('Times',20),bg='purple2').grid(row=4,column=0,sticky=W)
        Label(modify_main,text="New Food Name: ",font=('Times',20),bg='purple2').grid(row=5,column=0,sticky=W)
        Label(modify_main,text="New Quantity: ",font=('Times',20),bg='purple2').grid(row=6,column=0,sticky=W)
        Label(modify_main,text="New Price (in Rupees): ",font=('Times',20),bg='purple2').grid(row=7,column=0,sticky=W)

       
        Entry(modify_main,textvariable=e1,font=('Times',20)).grid(row=3,column=1,sticky=W)
        Entry(modify_main,textvariable=e2,font=('Times',20)).grid(row=4,column=1,sticky=W)
        Entry(modify_main,textvariable=e3,font=('Times',20)).grid(row=5,column=1,sticky=W)
        Entry(modify_main,textvariable=e4,font=('Times',20)).grid(row=6,column=1,sticky=W)
        Entry(modify_main,textvariable=e5,font=('Times',20)).grid(row=7,column=1,sticky=W)

        Button(modify_main,text="MODIFY FOOD",font=('Times',25),fg='blue2',bg='green',command=lambda: modify(somelist)).grid(row=8,column=2,sticky=W)
        modify_main.mainloop()
#______________________________________________________________________________________________________________________________________
#_________________________________________________Function to delete existing foods in the inventory______________________________________________
    def delete(self,wb,ws,ws1,ws2,inv):
        def delete_food():
            check=0;whatrow=0
            foodno=e1.get()
            foodname=e2.get()
            maxrow=ws.max_row
            if foodno != '' or foodname != '' :
                for i in range(2,maxrow+1):
                    if str(foodno).upper()==str(ws['b'+str(i)].value).upper() and str(foodname).upper()==str(ws['c'+str(i)].value).upper():
                        check=1
                        whatrow=i                   
                        break
                else:
                    tkinter.messagebox.showerror('Error','Food item does not exist. \n Enter a valid food item to delete')
            else:
                tkinter.messagebox.showerror('Error','Both fields are mandatory')
            if check==1:
                for i in range(1,maxrow+1):
                    if i == int(whatrow):
                        ws['a'+str(i)].value=''
                        ws['b'+str(i)].value=''
                        ws['c'+str(i)].value=''
                        ws['d'+str(i)].value=''
                        ws['e'+str(i)].value=''

                        ws1['a'+str(i)].value=''
                        ws1['b'+str(i)].value=''
                        ws1['c'+str(i)].value=''
                        ws1['d'+str(i)].value=''
                        ws1['e'+str(i)].value=''

                        ws2['a'+str(i)].value=''
                        ws2['b'+str(i)].value=''
                        ws2['c'+str(i)].value=''
                        ws2['d'+str(i)].value=''
                        ws2['e'+str(i)].value=''
                                                                   
                tkinter.messagebox.showinfo('Success','Food Item Deleted')
                                              
                wb.save('master_sheet.xlsx')
                delete.destroy()
                inv.destroy()
                self.inventory()
        delete=Toplevel()
        delete.geometry('600x200+100+100')
        delete.configure(bg='magenta')
        delete.title('The Food Box - Delete Existing Food Item')
        e1=StringVar()
        e2=StringVar()        
        Label(delete,text="Enter the details",font=('Times',30),bg='magenta').grid(row=0,column=0,columnspan=2,sticky=N+W+E+S)
        Label(delete,text="Food Number",font=('Times',20),bg='magenta').grid(row=1,column=0,sticky=W)
        Label(delete,text="Food Name",font=('Times',20),bg='magenta').grid(row=2,column=0,sticky=W)        
        Entry(delete,textvariable=e1,font=('Times',20)).grid(row=1,column=1,sticky=W)
        Entry(delete,textvariable=e2,font=('Times',20)).grid(row=2,column=1,sticky=W)
        Button(delete,text="DELETE",font=('Times',25),fg='blue2',bg='green',command=delete_food).grid(row=4,column=4,rowspan=2,columnspan=2,sticky=W)
        delete.mainloop()
#____________________________________________________________________________________________________________________________________________________________________
#_______________________________________________________Function to search food-items by their name_________________________________________________________________        
    def search_by_name(self,wb,ws,ws1,ws2): 
        def search1():
            foodname=e1.get()
            maxrow=ws.max_row
            Label(sbn,text="Food Type",font=('Times',20),bg='chocolate1').grid(row=8,column=0,sticky=W)
            Label(sbn,text="Food No",font=('Times',20),bg='chocolate1').grid(row=8,column=1,sticky=W)
            Label(sbn,text="Food Name",font=('Times',20),bg='chocolate1').grid(row=8,column=2,sticky=W)
            Label(sbn,text="Quantity",font=('Times',20),bg='chocolate1').grid(row=8,column=3,sticky=W)
            Label(sbn,text="Price",font=('Times',20),bg='chocolate1').grid(row=8,column=4,sticky=W)                
            if foodname != '':
                for i in range(2,maxrow+1):
                    if str(foodname).upper()==str(ws['c'+str(i)].value).upper():
                        l1=Label(sbn,text=str(ws['a'+str(i)].value).upper(),font=('Times',20),bg='chocolate1').grid(row=self.rowrow+5,column=0,sticky=W)
                        l2=Label(sbn,text=str(ws['b'+str(i)].value).upper(),font=('Times',20),bg='chocolate1').grid(row=self.rowrow+5,column=1,sticky=W)
                        l3=Label(sbn,text=str(ws['c'+str(i)].value),font=('Times',20),bg='chocolate1').grid(row=self.rowrow+5,column=2,sticky=W)
                        l4=Label(sbn,text=str(ws['d'+str(i)].value),font=('Times',20),bg='chocolate1').grid(row=self.rowrow+5,column=3,sticky=W,padx=10)
                        l5=Label(sbn,text='Rs. '+str(ws['e'+str(i)].value),font=('Times',20),bg='chocolate1').grid(row=self.rowrow+5,column=4,sticky=W,padx=10)
                        self.rowrow+=1
                        break
                
                else:
                    tkinter.messagebox.showinfo('Search Results','Food item does not exist')
                    sbn.destroy()
            else:
                tkinter.messagebox.showerror('Error','All fields are mandatory!')

        sbn=Toplevel()
        sbn.configure(bg='chocolate1')

        sbn.title('The Food Box - Search By Food Name')
        e1=StringVar()
        Label(sbn,text="Enter Food Name: ",font=('Times',20),bg='chocolate1').grid(row=1,column=0,sticky=W)       
        Entry(sbn,textvariable=e1,font=('Times',20)).grid(row=1,column=1,sticky=W)       
        b1=Button(sbn,text="Search for food",fg='blue2',bg='green',font=('Times',20),command=lambda: search1())
        b1.grid(row=6,column=1,padx=10,pady=5,sticky=W)        
        sbn.mainloop()
#_______________________________________________________________________________________________________________________________________________
#________________________________________________Function to search food-items by their food-numbers_______________________________________________________
        
    def search_by_fno(self,wb,ws,ws1,ws2):
         
        def search2():       
            foodno=e1.get()
            maxrow=ws.max_row
            Label(sbf,text="Food Type",font=('Times',20),bg='OrangeRed2').grid(row=3,column=0,sticky=W)
            Label(sbf,text="Food No",font=('Times',20),bg='OrangeRed2').grid(row=3,column=1,sticky=W)
            Label(sbf,text="Food Name",font=('Times',20),bg='OrangeRed2').grid(row=3,column=2,sticky=W)
            Label(sbf,text="Quantity",font=('Times',20),bg='OrangeRed2').grid(row=3,column=3,sticky=W)
            Label(sbf,text="Price",font=('Times',20),bg='OrangeRed2').grid(row=3,column=4,sticky=W)                    
             
            if foodno != '':
                for i in range(2,maxrow+1):
                    if str(foodno).upper()==str(ws['b'+str(i)].value).upper():
                        Label(sbf,text=str(ws['a'+str(i)].value),font=('Times',20),bg='OrangeRed2').grid(row= self.rowrow,column=0,sticky=W)
                        Label(sbf,text=str(ws['b'+str(i)].value),font=('Times',20),bg='OrangeRed2').grid(row= self.rowrow,column=1,sticky=W)
                        Label(sbf,text=str(ws['c'+str(i)].value),font=('Times',20),bg='OrangeRed2').grid(row= self.rowrow,column=2,sticky=W,padx=10)
                        Label(sbf,text=str(ws['d'+str(i)].value),font=('Times',20),bg='OrangeRed2').grid(row= self.rowrow,column=3,sticky=W,padx=10)
                        Label(sbf,text=str(ws['e'+str(i)].value),font=('Times',20),bg='OrangeRed2').grid(row= self.rowrow,column=4,sticky=W,padx=10)
                        self.rowrow+=1
                        break
                        
                        
                else:
                    tkinter.messagebox.showinfo('Search Results','Food item does not exist')
                    sbf.destroy()
                   
            else:
                tkinter.messagebox.showerror('Error','All fields are mandatory!')
        sbf=Toplevel()
        sbf.configure(bg='OrangeRed2')
        sbf.title('The Food Box - Search By Food Number')
        e1=StringVar()
        Label(sbf,text="Enter Food Number: ",font=('Times',20),bg='OrangeRed2').grid(row=1,column=0,sticky=W)        
        Entry(sbf,textvariable=e1,font=('Times',20)).grid(row=1,column=1,sticky=W)       
        Button(sbf,text="Search for food",font=('Times',20),fg='blue2',bg='green',command=lambda: search2()).grid(row=2,column=1,pady=5,padx=10,sticky=W)
        sbf.mainloop() 
#________________________________________________________________________________________________________________________________________________
#________________________________________________Function to generate reports after the sales of food-items_____________________________________________________

    def generate_report(self,wb,ws,ws1,ws2):
        def clear_report():
            for j in range(2,row+1):
                ws2['d'+str(j)]='0'
                wb.save('master_sheet.xlsx')
                wb.close()
            tkinter.messagebox.showinfo('The Food Box - Notification','Report has been reset for the day')
            generate_report_window.destroy()
        generate_report_window=Toplevel()
        day1=datetime.now()
        some_str=str(day1.day)+'/'+str(day1.month)+'/'+str(day1.year)
        generate_report_window.title('The Food Box - Report')
        generate_report_window.configure(bg='skyblue1')
        row=ws2.max_row
        Label(generate_report_window,text=some_str,bg='skyblue1',fg='black',font=('Times',20,'bold')).grid(row=0,column=2,sticky=N+E+W+S)
        for i in range(1,row+1):
            s1='a'+str(i);a1=str(ws2[s1].value)
            s2='b'+str(i);a2=str(ws2[s2].value)
            s3='c'+str(i);a3=str(ws2[s3].value)
            s4='d'+str(i);a4=str(ws2[s4].value)
            s5='e'+str(i);a5=str(ws2[s5].value)
            if a1!='None' and a2!='None' and a3!='None' and a4!='None' and a5!='None':
                Label(generate_report_window,text=a1.upper(),bg='skyblue1',fg='black',font=('Times',15,'bold')).grid(row=i,column=0,sticky=W)
                Label(generate_report_window,text=a2.upper(),bg='skyblue1',fg='black',font=('Times',15,'bold')).grid(row=i,column=1,sticky=W)
                Label(generate_report_window,text=a3.capitalize(),bg='skyblue1',fg='black',font=('Times',15,'bold')).grid(row=i,column=2,sticky=W)
                Label(generate_report_window,text=a4,bg='skyblue1',fg='black',font=('Times',15,'bold')).grid(row=i,column=3,sticky=W)
                Label(generate_report_window,text='Rs. '+a5,bg='skyblue1',fg='black',font=('Times',15,'bold')).grid(row=i,column=4,sticky=W)

        Button(generate_report_window,text='RESET REPORT',bg='orange',fg='black',font=('Times',15,'bold'),command=clear_report).grid(row=row+1,column=3,sticky=W)
                
                
        generate_report_window.mainloop()        
#________________________________________________________________________________________________________________________________________
#__________________________________________END OF THE INVENTORY CLASS MODULE______________________________________________________        


#$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$
#$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$


#__________________________________________________________________________________________________________________________________
#----------------------------------------Beginning of the sales module class--------------------------------------------------------------------------------

class Sales(Inventory):
 
    def __init__(self):
        self.value1=0
        self.wb=load_workbook('master_sheet.xlsx')
        self.ws_inv=self.wb['Inventory']
        self.ws_sales=self.wb['Sales']
        self.ws_report=self.wb['Report']
        self.dict1={}
#________________________________________________________________________________________________________________________________
    def save_data(self):
        self.wb.save('master_sheet1.xlsx')
        os.remove('master_sheet.xlsx')
        os.rename('master_sheet1.xlsx','master_sheet.xlsx')
#______________________________________________Function for the main sales page____________________________________________________
    def sales_module_user(self):
        def inc_and_dec(x,y,z,action):
            if action=='add':
                if x<int(self.ws_inv['d'+str(y)].value):
                    z['text']=str(x+1)
                    z.grid(row=y,column=10,sticky=W)                    
                else:
                   tkinter.messagebox.showerror("Error","This is the maximum stock in the inventory",parent=sales2)                              
            elif action=='remove':
                if int(z['text']) >=1:
                    z['text']=str(x-1)
                    z.grid(row=y,column=10,sticky=W)
                    
            self.dict1[y]=int(z['text'])
        sales2 = Toplevel()
        sales2.geometry('1500x1500')
        sales2.title('The Food Box - SALES')
        sales2.configure(bg='white')
        now=datetime.now()
        today1=str(now.day)+'/'+str(now.month)+'/'+str(now.year)
        img2=PhotoImage(file='csfb_logo.gif')
        Label(sales2,image=img2).grid(row=0,column=12)
        Label(sales2,text=today1,bg='white',fg='black',font=('Times',25,'bold'),width=8).grid(row=0,column=4,sticky=N+E+W+S)
        Label(sales2,text='Food Type',borderwidth=2,bg='yellow',relief='ridge',font=('Times',20)).grid(row=1,column=0,columnspan=2,sticky=W)
        Label(sales2,text='Food Number',borderwidth=2,bg='yellow',relief='ridge',font=('Times',20)).grid(row=1,column=2,columnspan=2,sticky=W)
        Label(sales2,text='Food Name',borderwidth=2,bg='yellow',relief='ridge',font=('Times',20)).grid(row=1,column=4,columnspan=3,sticky=W)
        Label(sales2,text='Price',borderwidth=2,bg='yellow',relief='ridge',font=('Times',20)).grid(row=1,column=7,sticky=W)
        Label(sales2,text='Order',borderwidth=2,bg='yellow',relief='ridge',font=('Times',20)).grid(row=1,column=9,columnspan=3,sticky=W)
        maxrow=self.ws_inv.max_row
        for i in range(2,maxrow+1):
            s1='a'+str(i);ftype=str(self.ws_inv[s1].value)
            s2='b'+str(i);fno=str(self.ws_inv[s2].value)
            s3='c'+str(i);fname=str(self.ws_inv[s3].value)
            s4='d'+str(i);qty=str(self.ws_inv[s4].value)
            s5='e'+str(i);price=str(self.ws_inv[s5].value)
            if qty != '0' and ftype!='None' and fno!='None' and fname!='None' and qty!='None' and price!='None':
                Label(sales2,text=ftype.upper(),font=('Times',14,"bold"),bg='white',fg='black').grid(row=i,column=0,columnspan=2,sticky=W)
                Label(sales2,text=fno.upper(),font=('Times',14,"bold"),bg='white',fg='black').grid(row=i,column=2,columnspan=2,sticky=W)
                Label(sales2,text=fname.capitalize(),font=('Times',14,"bold"),fg='black',bg='white').grid(row=i,column=4,columnspan=3,sticky=W)
                Label(sales2,text='Rs. '+price,font=('Times',14,"bold"),bg='white',fg='black').grid(row=i,column=7,sticky=W)
                l1=Label(sales2,text='0',font=('Times',16));l1.grid(row=i,column=10,sticky=W)
                Button(sales2,text='-',bg='orange red',font=('Times',15),command=lambda c=i,d=l1: inc_and_dec(int(d['text']),c,d,'remove')).grid(row=i,column=8,sticky=W)
                Button(sales2,text='+',bg='green2',font=('Times',15),command=lambda a=i,b=l1: inc_and_dec(int(b['text']),a,b,'add')).grid(row=i,column=11,sticky=W)
            self.dict1[i]=0
        Button(sales2,text="Proceed to Checkout",command=lambda: self.Checkout(sales2),bg='yellow',font=('Times',20)).grid(row=maxrow+1,column=2,padx=20,pady=20)
        sales2.mainloop()
#__________________________________________________________________________________________________________________________________________________________
#____________________________________________________WORKING FUNCTION__________________________________________________________________________________
    def Checkout(self,sales2):
        row=self.ws_sales.max_row; xz=0
        for i in range(2,row-1):
            if self.dict1[i]==0:
                xz+=1
        if xz==(row-3):
            tkinter.messagebox.showerror('Error','No Food Ordered!\nPlease Order Something!!!')
        else:
            for i in range(2,row+1):
                string='d'+str(i)
                try:
                    self.ws_sales[string]=self.dict1[i]
                except:
                    pass
            self.save_data()
            self.wb.close()
            self.lastpage(sales2)
#_____________________________________________________Function to display the bill of the selected items________________________________________________________
    def lastpage(self,sales2):
        lastpage=Tk()
        lastpage.title('Bill Payment')
        maxrow=self.ws_sales.max_row
        lastpage.configure(bg='ivory2')
        Label(lastpage,text='FOOD TYPE',font=('courier new',20,'bold'),borderwidth=2,relief='ridge',bg='ivory2').grid(row=0,column=0,sticky=W)
        Label(lastpage,text='FOOD NUMBER',font=('courier new',20,'bold'),borderwidth=2,relief='ridge',bg='ivory2').grid(row=0,column=1,sticky=W)
        Label(lastpage,text='FOOD NAME',font=('courier new',20,'bold'),borderwidth=2,relief='ridge',bg='ivory2').grid(row=0,column=2,sticky=W)
        Label(lastpage,text='QUANTITY',font=('courier new',20,'bold'),borderwidth=2,relief='ridge',bg='ivory2').grid(row=0,column=3,sticky=W)
        Label(lastpage,text=' TOTAL PRICE',font=('courier new',20,'bold'),borderwidth=2,relief='ridge',bg='ivory2').grid(row=0,column=4,sticky=W,padx=20)
        count=0
        for i in range(2,maxrow+1):
            s1='a'+str(i);ftype=str(self.ws_sales[s1].value)
            s2='b'+str(i);fno=str(self.ws_sales[s2].value)
            s3='c'+str(i);fname=str(self.ws_sales[s3].value)
            s4='d'+str(i);order=str(self.ws_sales[s4].value)
            s5='e'+str(i);price=str(self.ws_sales[s5].value)
            string ='d'+ str(i)
            try:
                if int(self.ws_sales[string].value)!=0:
                    Label(lastpage,text=ftype.upper(),font=('courier new',20,'bold'),bg='ivory2').grid(row=i,column=0,sticky=W)
                    Label(lastpage,text=fno.upper(),font=('courier new',20,'bold'),bg='ivory2').grid(row=i,column=1,sticky=W)
                    Label(lastpage,text=fname.capitalize(),font=('courier new',20,'bold'),bg='ivory2').grid(row=i,column=2,sticky=W)
                    Label(lastpage,text=order,font=('courier new',20,'bold'),bg='ivory2').grid(row=i,column=3,sticky=N+E+W+S)
                    Label(lastpage,text='Rs. '+str(int(order)*int(price)),font=('courier new',20,'bold'),bg='ivory2').grid(row=i,column=4,sticky=W,padx=20)
                    count+=int(order)*int(price)
            except:
                pass
        def pay_bill():
            self.report(sales2)
            lastpage.destroy()

        Label(lastpage,text="GROSS TOTAL: Rs. "+ str(count),bg='black',fg='green',font=('courier new',20,'bold')).grid(row=maxrow+1,column=4)
        Button(lastpage,text="GO BACK AND ORDER MORE FOOD",font=('Times',20,'bold'),bg='red',command=lastpage.destroy).grid(row=maxrow+2,column=0,columnspan=2)
        Button(lastpage,text="PAY BILL AND ENJOY YOUR MEAL",font=('Times',20,'bold'),bg='green',command=pay_bill).grid(row=maxrow+2,column=3,columnspan=2)
#____________________________________________________________________________________________________________________________________________________________
#____________________________________________________________Function to take the user to the last page___________________________________________________
        
    def report(self,sales2):
        def gotomainpage():
            sales2.destroy()
            thankyou.destroy()
        maxrow=self.ws_report.max_row
        for i in range(2,maxrow+1):
            try:
                Ordered_food=int(self.ws_sales['d'+str(i)].value)
                self.ws_inv['d'+str(i)].value=str(int(self.ws_inv['d'+str(i)].value)-Ordered_food)
                self.ws_report['d'+str(i)].value=str(int(self.ws_report['d'+str(i)].value)+Ordered_food)
                self.ws_sales['d'+str(i)].value=0
            except:
                pass
          
        thankyou=Toplevel()
        thankyou.title('The Food Box')
        thankyou.configure(bg='firebrick1')
        thankyou.geometry("500x250+250+125")
        Label(thankyou,text="Thank You For",font=('Times',40),bg="yellow",fg="red").grid(row=0,column=0)
        Label(thankyou,text="Using The Food Box",font=('Times',40),bg="yellow",fg="red").grid(row=1,column=0)
        Label(thankyou,text="Enjoy your meal!",font=('Times',40),bg="yellow",fg="red").grid(row=2,column=0)
        self.wb.save('master_sheet.xlsx')
        thankyou.protocol('WM_DELETE_WINDOW', gotomainpage)
#__________________________________________________________________________________________________________________________________
#______________________________________________END OF THE SALES CLASS MODULE_________________________________________________      
        

#$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$
#$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$

#__________________________________________________________________________________________________________________________________
#----------------------------------------Beginning of the MAIN PROGRAM--------------------------------------------------------------------------------

#__________________________________________________________________________________________________________________________________
#________________________________________Login page function________________________________________________________________________

def login_page(x):
    def gotopage(somenum):
        if somenum==0:
            access_sales=Sales()
            access_sales.sales_module_user()
        else:
            access_inventory=Inventory()
            access_inventory.inventory()
            
    def validate():
        username=a.get()
        password=b.get()
        wb=load_workbook('master_sheet.xlsx')
        if x==1:
            ws1=wb['Admin']; row1=ws1.max_row
            for i in range(1,row1+1):
                string='b'+str(i); string1='c'+str(i)
                if (ws1[string].value==username) and (ws1[string1].value==password):
                    root_login.destroy()
                    gotopage(1)
                    break
                    #ADMIN LOGIN INTO INVENTORY
            else:
                tkinter.messagebox.showerror('Error','Wrong username or password',parent=root_login)
        elif x==0:
            ws1=wb['Admin']; row1=ws1.max_row
            ws2=wb['User']; row2=ws2.max_row
            goto='';abc=0
            for i in range(1,row1+1):
                string='b'+str(i); string1='c'+str(i)
                if (ws1[string].value==username) and (ws1[string1].value==password):
                    root_login.destroy()
                    goto='smu'
                    abc=1
                    break
                #ADMIN LOGIN INTO SALES
            if abc != 1:
                for j in range(1,row2+1):
                    string='b'+str(j); string1='c'+str(j)
                    if (ws2[string].value==username) and (ws2[string1].value==password):
                        root_login.destroy()
                        goto='smu'
                        abc=1
                        break
                    #USER LOGIN INTO SALES
            if abc==1 and goto=='smu' or goto=='sma':
                gotopage(0)
            elif abc != 1:
                tkinter.messagebox.showerror('Error','Wrong username or password',parent=root_login)
        wb.close()
    root_login=Toplevel()
    root_login.geometry('1050x500')
    w = 1050
    h = 500
    ws = root_login.winfo_screenwidth()
    hs = root_login.winfo_screenheight()
    x1 = (ws/2) - (w/2)
    y1 = (hs/2) - (h/2)
    root_login.geometry('%dx%d+%d+%d' % (w, h, x1, y1))
    root_login.maxsize(height=500,width=1050)
    root_login.configure(bg='gold')

    a=StringVar()
    b=StringVar()

    img=PhotoImage(file='csfb_main.gif')
    Label(root_login,image=img,bg='gold',height=400,width=600).grid(row=0,column=0,rowspan=30)

    Label(root_login,text='Username: ',font=('Times',20),bg='gold',fg='blue2').grid(row=20,column=2,sticky=W)
    Entry(root_login,font=('Times',20),textvariable=a).grid(row=20,column=3)
    Label(root_login,text='Password: ',font=('Times',20),bg='gold',fg='blue2').grid(row=21,column=2,sticky=W)
    Entry(root_login,font=('Times',20),show='*',textvariable=b).grid(row=21,column=3)

    Button(root_login,text='LOGIN',bg='medium blue',fg='gold',height=3,width=15,command=validate).grid(row=24,column=3)

    root_login.mainloop()

#________________________________________________________________________________________________________________________________________
#_________________________________________Function to facilitate sign-up of new users_______________________________________________________

def sign_up_now():
    root1=Toplevel(root)
    root1.configure(bg='DeepSkyBlue2')
    w = 650
    h = 330
    ws = root1.winfo_screenwidth()
    hs = root1.winfo_screenheight()
    x2 = (ws/2) - (w/2)
    y2 = (hs/2) - (h/2)
    root1.geometry('%dx%d+%d+%d' % (w, h, x2, y2))
    root1.maxsize(height=330,width=650)

    user1=StringVar()
    pass1=StringVar()
    pass2=StringVar()
    name1=StringVar()
    v=IntVar()

    def signup():
        name=name1.get()
        user=user1.get()
        pass_1=pass1.get()
        pass_2=pass2.get()
        person=v.get()
        
        def add_user(name,user,pass_1):
            wb=load_workbook('master_sheet.xlsx')
            ws=wb['User']
            ws.append([name,user,pass_1])
            wb.save('master_sheet1.xlsx')
            os.remove('master_sheet.xlsx')
            os.rename('master_sheet1.xlsx','master_sheet.xlsx')
            wb.close()
        def add_admin(name,user,pass_1):
            wb=load_workbook('master_sheet.xlsx')
            ws=wb['Admin']
            ws.append([name,user,pass_1])
            wb.save('master_sheet1.xlsx')
            os.remove('master_sheet.xlsx')
            os.rename('master_sheet1.xlsx','master_sheet.xlsx')
            wb.close()
        str1=string.ascii_letters
        str2=string.digits
        str3=string.punctuation
        lst1=[];lst2=[];lst3=[]
        if (len(name)>0 and len(user)>0):
            if len(pass_1)>=8:
                for i in pass_1:
                    if i in str1:
                        lst1.append(i)
                    elif i in str2:
                        lst2.append(i)
                    elif i in str3:
                        lst3.append(i)
                if len(lst1)>0 and len(lst2)>0 and len(lst3)>0:
                    if (pass_1==pass_2) and person==0:
                        add_user(name,user,pass_1)
                        tkinter.messagebox.showinfo('Success','Signup successful!',parent=root1)
                        root1.destroy()
                    elif (pass_1==pass_2) and person==1:
                        add_admin(name,user,pass_1)
                        tkinter.messagebox.showinfo('Success','Signup successful!',parent=root1)
                        root1.destroy()
                    else:
                        tkinter.messagebox.showerror('Error','Passwords don\'t match',parent=root1)
                else:
                    tkinter.messagebox.showerror('Error','Password must contain combination of\nletters,digits and special characters',parent=root1)
            else:
                tkinter.messagebox.showerror('Error','Password must be atleast 8 characters long',parent=root1)
        else:
            tkinter.messagebox.showerror('Error','No name or username given',parent=root1)
    Label(root1,text='WELCOME NEW USER',justify=CENTER,font=('Times',30),fg='gray1',bg='DeepSkyBlue2').grid(row=0,column=0,rowspan=2,columnspan=3,sticky=N+E+W+S)

    Label(root1,text='Enter your name: ',font=('Times',20),bg='DeepSkyBlue2',fg='gray1').grid(row=3,column=0,sticky=W)
    Entry(root1,font=('Times',20),textvariable=name1,bg='white').grid(row=3,column=1)
    Label(root1,text='Enter your preferred username: ',font=('Times',20),bg='DeepSkyBlue2',fg='gray1').grid(row=4,column=0,sticky=W)
    Entry(root1,font=('Times',20),textvariable=user1,bg='white').grid(row=4,column=1)
    Label(root1,text='Enter password: ',font=('Times',20),bg='DeepSkyBlue2',fg='gray1').grid(row=5,column=0,sticky=W)
    Entry(root1,font=('Times',20),show='*',textvariable=pass1,bg='white').grid(row=5,column=1)
    Label(root1,font=('Times',20),text='Confirm password: ',bg='DeepSkyBlue2',fg='gray1').grid(row=6,column=0,rowspan=2,sticky=W)
    Entry(root1,font=('Times',20),show='*',textvariable=pass2,bg='white').grid(row=6,column=1,rowspan=2)

    Button(root1,text='SIGN UP',fg='yellow2',bg='red2',height=3,width=15,command=signup).grid(row=16,column=1,rowspan=2,pady=4)

    Radiobutton(root1,text='I am an Administrator',font=('Times',10),bg='DeepSkyBlue2',fg='blue',variable=v,value=1).grid(row=13,column=1,pady=4,sticky=W)
    Radiobutton(root1,text='I am a User',font=('Times',10),bg='DeepSkyBlue2',fg='blue',variable=v,value=0).grid(row=10,column=1,pady=4,sticky=W)

    root1.mainloop()

#$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$
#$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$

#________________________________________*************************MAIN PAGE*********************************_____________________________________

root=Tk()
#-------------------Positioning of the root-----------------------------
w = 1000
h = 600
ws = root.winfo_screenwidth()
hs = root.winfo_screenheight()
x3 = (ws/2) - (w/2)
y3 = (hs/2) - (h/2)
root.geometry('%dx%d+%d+%d' % (w, h, x3, y3))
root.maxsize(height=600,width=1000)
#--------------------------------------------------------------------------
root.title('Welcome to the Food Box')
root.configure(bg='blue2')

img=PhotoImage(file='welcome-food1.gif')
Label(root,image=img,bg='blue2',height=600,width=600).grid(row=0,column=0,rowspan=10)

Button(root,text='Sales',font=('Times',20),fg='black',bg='firebrick1',height=5,width=25,command=lambda: login_page(x=0)).grid(row=2,column=1)
Button(root,text='Inventory\n(Only Administrator Access)',fg='black',bg='gold',font=('Times',20),height=5,width=25,command=lambda: login_page(x=1)).grid(row=3,column=1)
Button(root,text='New to Food Box?\nSign Up Now!',font=('Times',20),fg='black',bg='green',height=5,width=25,command=sign_up_now).grid(row=4,column=1)

root.mainloop()

#_________________________________________________________________________________________________________________________________________
#************************************************************END OF THE MAIN PROGRAM****************************************************************************
